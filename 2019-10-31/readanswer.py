

# The next script we're looking at needs to be able to read all of the excel files in a folder (likely 500+) 
# taking certain cell values and populating two other excel sheets. In the answer spreadsheet row 2 and 3 
# describe where to get the data from in the question spreadsheet

'''
2010-10-31  
1. The script that reads the questions files can we put the question file name on each row in column c of the answers file
2. The second is that file in another format 
'''

import os
import openpyxl
# import pdb
import warnings
from collections import namedtuple
from string import digits

def read_q_file_and_fill_ans(q_file, sheets_cells, sh_ans, first_empty_column):

	book_q = openpyxl.load_workbook(q_file)
	for sheet in sheets_cells:
		try:
			q_sheet = book_q[sheet]
		except KeyError as e:
			print("Sheet {} not found in question file {}.".format(sheet, q_file))
			continue
		for q_cell, row in sheets_cells[sheet]:
			if q_cell:
				try:
					sh_ans.cell(row, first_empty_column).value = q_sheet[q_cell].value
				except ValueError as e:
					print(e)
				except TypeError as e:
					# pdb.set_trace()
					print(e)
				except AttributeError as e:
					pdb.set_trace()
	book_q.close()

def get_column_range(sh, range_cells):

	return [sh[i].column for i in range_cells.split(":")]

def check_merge_cell(cell_coordinate):
	if '/' in cell_coordinate:
		first, second = cell_coordinate.split('/')
		cell_coordinate = first + ''.join([i for i in second if i in digits])

	return cell_coordinate

def extract_letter_from_coordinate(coordinate):
	letters = ''.join(l for l in coordinate if l not in digits)
	return letters

def find_first_empty_column(sh_ans, first_empty_row, first_full_column):
	col = first_full_column
	while sh_ans.cell(first_empty_row, col).value:
		col += 1
	return col

def read_fill_ans_file(answer_file, q_files, main_parameters):

	sheets_cells = {}
	sh_ans = None

	if not os.path.isfile(answer_file):
		print("Not found file {}".format(answer_file))
		return

	book_ans = openpyxl.load_workbook(answer_file)

	for sheet_in_answer in main_parameters.sheetnames_in_answer:

		try:
			sh_ans = book_ans[sheet_in_answer]
		except KeyError as e:
			# print("Sheet {} not not found. Error:{}".format(sheet_in_answer, e))
			continue
		
		for k in main_parameters.describe_cell:
			if sh_ans.cell(*k).value != main_parameters.describe_cell[k]:
				print('Not found cell"{}" which describe where to get data  \
					from the question spreadsheet'.format(main_parameters.describe_cell[k]))
				break

			if main_parameters.describe_cell[k] == "Sheet":
				first_empty_row = k[0]
				first_full_column = k[1]
				for i in range(k[0] + 1, sh_ans.max_row + 1):
					sheet_name = sh_ans.cell(i, k[1]).value
					cell_coordinate = sh_ans.cell(i, k[1] + 1).value

					if sheet_name and cell_coordinate:
						cell_coordinate = check_merge_cell(cell_coordinate)

						if  sheet_name in sheets_cells:
							sheets_cells[sheet_name].append((cell_coordinate, i))
						else:
							sheets_cells[sheet_name] = [(cell_coordinate, i)]

		if sh_ans:
			first_empty_column = find_first_empty_column(sh_ans, first_empty_row + 1, first_full_column)
			# pdb.set_trace()
			for q_file in q_files:
				full_path_q_file = os.path.join(main_parameters.dir_question_files, q_file)
				sh_ans.cell(first_empty_row, first_empty_column).value = q_file
				print("Question file: {}".format(full_path_q_file))
				read_q_file_and_fill_ans(full_path_q_file, sheets_cells, sh_ans, first_empty_column)
				first_empty_column += 1

		book_ans.save(answer_file)
		book_ans.close()

						
def main():
	warnings.filterwarnings("ignore")

	Parameters = namedtuple('Parameters', [
		'dir_question_files',
		'dir_answer_files',
		'describe_cell',
		'sheetnames_in_answer'
		])	

	main_parameters = Parameters(
		dir_question_files = 'questions',
		dir_answer_files = 'answers',
		describe_cell = {(11,2): 'Sheet', (11,3): 'Cell'},
		sheetnames_in_answer = ['Evaluation', 'Admin', 'Evaluation Neville']
	)

	q_files = [f for f in os.listdir(main_parameters.dir_question_files) if f.endswith('xlsx')]
	answer_files = [os.path.join(main_parameters.dir_answer_files, f) 
					 for f in os.listdir(main_parameters.dir_answer_files) if f.endswith('xlsx')]
	
	for ans_file in answer_files:
		print("Answer file: {}".format(ans_file))
		read_fill_ans_file(ans_file, q_files, main_parameters)

if __name__ == "__main__":
	main()
