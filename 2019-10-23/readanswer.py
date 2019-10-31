# The next script we're looking at needs to be able to read all of the excel files in a folder (likely 500+) 
# taking certain cell values and populating two other excel sheets. In the answer spreadsheet row 2 and 3 
# describe where to get the data from in the question spreadsheet

import os
import openpyxl
import string
import pdb
import warnings


def read_and_fill(q_file, sheets_q, sh_ans, first_empty_row):

	book_q = openpyxl.load_workbook(q_file)
	for sheet in sheets_q:
		try:
			q_sheet = book_q[sheet]
		except KeyError as e:
			print("Sheet {} not not found. Error:{}".format(sheet, e))
			continue
		for q_cell, column in sheets_q[sheet]:
			if q_cell:
				try:
					sh_ans.cell(first_empty_row, column).value = q_sheet[q_cell].value
				except ValueError as e:
					print(e)
				except TypeError as e:
					# pdb.set_trace()
					print(e)

	book_q.close()

def get_column_range(sh, range_cells):

	return [sh[i].column for i in range_cells.split(":")]


def read_ans_file(answer_file, dir_question_files, q_files, describe_cell, sheetnames_in_answer):

	sheets_q = {}

	book_ans = openpyxl.load_workbook(answer_file)

	for sheet_in_answer in sheetnames_in_answer:

		try:
			sh_ans = book_ans[sheet_in_answer]
		except KeyError as e:
			# print("Sheet {} not not found. Error:{}".format(sheet_in_answer, e))
			continue
		
		for k in describe_cell:
			if sh_ans.cell(*k).value != describe_cell[k]:
				print('Not found cell"{}", \
					which describe where to get data from the question spreadsheet'.format(describe_cell[k]))
				break

			if describe_cell[k] == "Sheet":
				merged_cells = [i.coord for i in sh_ans.merged_cells.ranges if str(k[0]) in i.coord]
				for i in range(k[1] + 1, sh_ans.max_column + 1):
					cell = sh_ans.cell(k[0], i)
					if cell.value:
						sheet_name = cell.value.strip()

						range_cells = [i for i in merged_cells if i.startswith(cell.coordinate)]
						if range_cells:
							c_range = get_column_range(sh_ans, range_cells[0])
							sheets_q[sheet_name]= [(sh_ans.cell(cell.row + 1, i).value, i) 
														for i in range(c_range[0], c_range[1] + 1)]
						else:
							sheets_q[sheet_name] = [(sh_ans.cell(cell.row + 1, cell.column).value, cell.column)]
	# pdb.set_trace()
	first_empty_row = sh_ans.max_row + 1
	for i, q_file in enumerate(q_files):
		print("Question file: {}".format(q_file))
		read_and_fill(os.path.join(dir_question_files, q_file), 
			sheets_q, sh_ans, first_empty_row)
		first_empty_row += 1

	book_ans.save(answer_file)
	book_ans.close()

def main():
	warnings.filterwarnings("ignore")

	dir_question_files = "questions"
	q_files = [f for f in os.listdir(dir_question_files) if f.endswith('xlsx')]
	
	dir_answers_files = "answers"
	answer_files = [f for f in os.listdir(dir_answers_files) if f.endswith('xlsx')]

	describe_cell = {(2,3): 'Sheet', (3,3): 'Cell'}

	sheetnames_in_answer = ['Evaluation', 'Admin']

	for ans_file in answer_files:
		print("Answer file: {}".format(ans_file))
		read_ans_file(os.path.join(dir_answers_files, ans_file), 
									dir_question_files, q_files, describe_cell, sheetnames_in_answer)

if __name__ == "__main__":
	main()

